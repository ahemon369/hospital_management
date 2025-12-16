from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from patients.models import Patient
from appointments.models import Appointment
from billing.models import Bill
from pharmacy.models import Medicine

@login_required
def reports_home(request):
    """Reports dashboard with overview statistics"""
    total_appointments = Appointment.objects.count()
    total_revenue = Bill.objects.filter(status='PAID').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    total_patients = Patient.objects.count()
    medicines_in_stock = Medicine.objects.filter(stock_quantity__gt=0).count()
    
    context = {
        'total_appointments': total_appointments,
        'total_revenue': total_revenue,
        'total_patients': total_patients,
        'medicines_in_stock': medicines_in_stock,
    }
    return render(request, 'reports/reports_home.html', context)


@login_required
def daily_appointments_report(request):
    """Daily appointments report"""
    from doctors.models import Doctor
    
    # Get date from request or use today
    date_str = request.GET.get('date')
    if date_str:
        from datetime import datetime
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        report_date = timezone.now().date()
    
    # Get appointments for the selected date
    appointments = Appointment.objects.filter(
        appointment_date=report_date
    ).select_related('patient__user', 'doctor__user').order_by('appointment_time')
    
    # Statistics
    total_today = appointments.count()
    pending = appointments.filter(status='PENDING').count()
    confirmed = appointments.filter(status='CONFIRMED').count()
    completed = appointments.filter(status='COMPLETED').count()
    cancelled = appointments.filter(status='CANCELLED').count()
    
    # Doctor statistics for this day
    doctor_stats = []
    doctors_today = appointments.values_list('doctor', flat=True).distinct()
    
    for doctor_id in doctors_today:
        try:
            doctor = Doctor.objects.select_related('user').get(pk=doctor_id)
            total = appointments.filter(doctor=doctor).count()
            doctor_stats.append({
                'doctor': doctor,
                'total': total,
            })
        except Doctor.DoesNotExist:
            continue
    
    # Sort by total
    doctor_stats = sorted(doctor_stats, key=lambda x: x['total'], reverse=True)
    
    context = {
        'appointments': appointments,
        'total_today': total_today,
        'pending': pending,
        'confirmed': confirmed,
        'completed': completed,
        'cancelled': cancelled,
        'report_date': report_date,
        'doctor_stats': doctor_stats,
    }
    return render(request, 'reports/daily_appointments_report.html', context)


@login_required
def daily_billing_report(request):
    """Daily billing report with real data"""
    # Get date from request or use today
    date_str = request.GET.get('date')
    if date_str:
        from datetime import datetime
        report_date = datetime.strptime(date_str, '%Y-%m-d').date()
    else:
        report_date = timezone.now().date()
    
    # Get today's bills
    bills = Bill.objects.filter(
        created_at__date=report_date
    ).select_related('patient__user').order_by('-created_at')
    
    # Calculate statistics
    total_bills = bills.count()
    total_amount = bills.aggregate(total=Sum('total_amount'))['total'] or 0
    paid_amount = bills.aggregate(total=Sum('amount_paid'))['total'] or 0
    pending_amount = total_amount - paid_amount
    
    # Status counts
    paid_count = bills.filter(status='PAID').count()
    unpaid_count = bills.filter(status='UNPAID').count()
    partial_count = bills.filter(status='PARTIAL').count()
    partial_amount = bills.filter(status='PARTIAL').aggregate(total=Sum('amount_paid'))['total'] or 0
    
    # Revenue breakdown
    consultation_total = bills.aggregate(total=Sum('consultation_fee'))['total'] or 0
    medicine_total = bills.aggregate(total=Sum('medicine_charges'))['total'] or 0
    lab_total = bills.aggregate(total=Sum('lab_charges'))['total'] or 0
    other_total = bills.aggregate(total=Sum('other_charges'))['total'] or 0
    
    context = {
        'bills': bills,
        'total_bills': total_bills,
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'pending_amount': pending_amount,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'partial_count': partial_count,
        'partial_amount': partial_amount,
        'consultation_total': consultation_total,
        'medicine_total': medicine_total,
        'lab_total': lab_total,
        'other_total': other_total,
        'report_date': report_date,
    }
    return render(request, 'reports/daily_billing_report.html', context)


@login_required
def medicine_stock_report(request):
    """Medicine stock report with real data"""
    
    # Get all medicines
    medicines = Medicine.objects.all()
    
    # Total medicines count
    total_medicines = medicines.count()
    
    # In stock (stock > 0)
    in_stock = medicines.filter(stock_quantity__gt=0).count()
    
    # Low stock (stock between 1 and 10)
    low_stock = medicines.filter(stock_quantity__gt=0, stock_quantity__lte=10).count()
    
    # Out of stock (stock = 0)
    out_of_stock = medicines.filter(stock_quantity=0).count()
    
    # Stock value calculations
    total_purchase_value = 0
    total_selling_value = 0
    
    for medicine in medicines:
        if medicine.stock_quantity and medicine.purchase_price:
            total_purchase_value += medicine.stock_quantity * medicine.purchase_price
        if medicine.stock_quantity and medicine.selling_price:
            total_selling_value += medicine.stock_quantity * medicine.selling_price
    
    expected_profit = total_selling_value - total_purchase_value
    profit_margin = (expected_profit / total_purchase_value * 100) if total_purchase_value > 0 else 0
    
    # Stock distribution
    adequate_stock = medicines.filter(stock_quantity__gt=50).count()
    moderate_stock = medicines.filter(stock_quantity__gte=11, stock_quantity__lte=50).count()
    critical_stock = medicines.filter(stock_quantity__lte=10).count()
    
    # Critical alerts (out of stock or low stock)
    critical_medicines = medicines.filter(
        Q(stock_quantity=0) | Q(stock_quantity__lte=10)
    ).order_by('stock_quantity')
    
    # Expiring soon (within 30 days)
    thirty_days_from_now = timezone.now().date() + timedelta(days=30)
    expiring_medicines = medicines.filter(
        expiry_date__lte=thirty_days_from_now,
        expiry_date__gte=timezone.now().date()
    ).order_by('expiry_date')
    
    # All medicines for complete inventory
    all_medicines = medicines.order_by('name')
    
    context = {
        # Summary stats
        'total_medicines': total_medicines,
        'in_stock': in_stock,
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        
        # Stock value
        'total_purchase_value': round(total_purchase_value, 2),
        'total_selling_value': round(total_selling_value, 2),
        'expected_profit': round(expected_profit, 2),
        'profit_margin': round(profit_margin, 2),
        
        # Distribution
        'adequate_stock': adequate_stock,
        'moderate_stock': moderate_stock,
        'critical_stock': critical_stock,
        'adequate_percent': round((adequate_stock / total_medicines * 100), 0) if total_medicines > 0 else 0,
        'moderate_percent': round((moderate_stock / total_medicines * 100), 0) if total_medicines > 0 else 0,
        'critical_percent': round((critical_stock / total_medicines * 100), 0) if total_medicines > 0 else 0,
        
        # Lists
        'critical_medicines': critical_medicines[:10],  # Top 10 critical
        'expiring_medicines': expiring_medicines[:10],  # Top 10 expiring
        'all_medicines': all_medicines,
    }
    
    return render(request, 'reports/medicine_stock_report.html', context)


@login_required
def doctor_appointments_report(request):
    """Doctor appointments report with real data"""
    from doctors.models import Doctor
    
    # Get date range from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    # Get all doctors with appointment counts
    doctors = Doctor.objects.all().select_related('user')
    
    doctor_stats = []
    total_appointments = 0
    total_completed = 0
    
    for doctor in doctors:
        # Filter appointments by date range
        appointments = Appointment.objects.filter(
            doctor=doctor,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        )
        
        total = appointments.count()
        completed = appointments.filter(status='COMPLETED').count()
        pending = appointments.filter(status='PENDING').count()
        confirmed = appointments.filter(status='CONFIRMED').count()
        cancelled = appointments.filter(status='CANCELLED').count()
        
        # Calculate completion rate
        completion_rate = (completed / total * 100) if total > 0 else 0
        
        doctor_stats.append({
            'doctor': doctor,
            'total_appointments': total,
            'completed': completed,
            'pending': pending,
            'confirmed': confirmed,
            'cancelled': cancelled,
            'completion_rate': round(completion_rate, 1),
        })
        
        total_appointments += total
        total_completed += completed
    
    # Sort by total appointments
    doctor_stats = sorted(
        doctor_stats, 
        key=lambda x: x['total_appointments'], 
        reverse=True
    )
    
    # Calculate overall stats
    active_doctors = len([d for d in doctor_stats if d['total_appointments'] > 0])
    avg_completion_rate = (total_completed / total_appointments * 100) if total_appointments > 0 else 0
    
    # Top 3 doctors
    top_doctors = sorted(
        [d for d in doctor_stats if d['total_appointments'] > 0],
        key=lambda x: x['completion_rate'],
        reverse=True
    )[:3]
    
    context = {
        'doctor_stats': doctor_stats,
        'active_doctors': active_doctors,
        'total_appointments': total_appointments,
        'total_completed': total_completed,
        'avg_completion_rate': round(avg_completion_rate, 1),
        'top_doctors': top_doctors,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/doctor_appointments_report.html', context)


@login_required
def export_report_pdf(request, report_type):
    """Export report as PDF using ReportLab"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        # Container for elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        if report_type == 'medicine_stock':
            # Title
            elements.append(Paragraph("MEDICINE STOCK REPORT", title_style))
            elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Get data
            medicines = Medicine.objects.all()
            total_medicines = medicines.count()
            in_stock = medicines.filter(stock_quantity__gt=0).count()
            out_of_stock = medicines.filter(stock_quantity=0).count()
            
            # Summary
            elements.append(Paragraph("SUMMARY", heading_style))
            summary_data = [
                ['Total Medicines', str(total_medicines)],
                ['In Stock', str(in_stock)],
                ['Out of Stock', str(out_of_stock)],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f9ff')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bfdbfe'))
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Medicine List
            elements.append(Paragraph("COMPLETE INVENTORY", heading_style))
            
            # Table headers
            data = [['ID', 'Name', 'Category', 'Stock', 'Price', 'Status']]
            
            # Table data
            for medicine in medicines[:50]:  # First 50 medicines
                status = 'In Stock' if medicine.stock_quantity > 0 else 'Out'
                data.append([
                    str(medicine.medicine_id)[:10],
                    str(medicine.name)[:25],
                    str(medicine.category)[:15],
                    str(medicine.stock_quantity),
                    f'৳{medicine.selling_price}',
                    status
                ])
            
            # Create table
            table = Table(data, colWidths=[1*inch, 2*inch, 1.2*inch, 0.8*inch, 1*inch, 0.8*inch])
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
            ]))
            
            elements.append(table)
            
            # Note
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("<i>Note: This report shows the first 50 medicines. For complete list, use Excel export.</i>", styles['Normal']))
        
        elif report_type == 'daily_appointments':
            # Title
            elements.append(Paragraph("DAILY APPOINTMENTS REPORT", title_style))
            
            today = timezone.now().date()
            elements.append(Paragraph(f"Date: {today.strftime('%B %d, %Y')}", styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Get appointments
            appointments = Appointment.objects.filter(
                appointment_date=today
            ).select_related('patient__user', 'doctor__user')
            
            # Summary
            elements.append(Paragraph("SUMMARY", heading_style))
            total = appointments.count()
            completed = appointments.filter(status='COMPLETED').count()
            pending = appointments.filter(status='PENDING').count()
            
            summary_data = [
                ['Total Appointments', str(total)],
                ['Completed', str(completed)],
                ['Pending', str(pending)],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f9ff')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bfdbfe'))
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Appointments list
            if appointments.exists():
                elements.append(Paragraph("APPOINTMENTS", heading_style))
                
                data = [['Time', 'Patient', 'Doctor', 'Type', 'Status']]
                
                for apt in appointments:
                    data.append([
                        apt.appointment_time.strftime('%I:%M %p'),
                        apt.patient.user.get_full_name()[:25],
                        f"Dr. {apt.doctor.user.get_full_name()}"[:25],
                        apt.appointment_type[:15],
                        apt.get_status_display()
                    ])
                
                table = Table(data, colWidths=[1.2*inch, 2*inch, 2*inch, 1.3*inch, 1*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("No appointments scheduled for this date.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Return response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
        response.write(pdf)
        return response
        
    except ImportError:
        # If ReportLab not installed
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_error.txt"'
        response.write('PDF export requires ReportLab library.\n')
        response.write('Install it using: pip install reportlab\n')
        return response
    except Exception as e:
        # Any other error
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_error.txt"'
        response.write(f'Error generating PDF: {str(e)}\n')
        return response


@login_required
def export_report_excel(request, report_type):
    """Export report as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()
    
    # Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Add headers based on report type
    if report_type == 'medicine_stock':
        medicines = Medicine.objects.all()
        
        # Add title
        ws['A1'] = 'MEDICINE STOCK REPORT'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:I1')
        
        # Add date
        ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Add headers
        headers = ['ID', 'Name', 'Category', 'Manufacturer', 'Stock', 
                  'Purchase Price', 'Selling Price', 'Expiry', 'Status']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, medicine in enumerate(medicines, 5):
            ws.cell(row=row_num, column=1, value=medicine.medicine_id)
            ws.cell(row=row_num, column=2, value=medicine.name)
            ws.cell(row=row_num, column=3, value=medicine.category)
            ws.cell(row=row_num, column=4, value=medicine.manufacturer or 'N/A')
            ws.cell(row=row_num, column=5, value=medicine.stock_quantity)
            ws.cell(row=row_num, column=6, value=float(medicine.purchase_price) if medicine.purchase_price else 0)
            ws.cell(row=row_num, column=7, value=float(medicine.selling_price) if medicine.selling_price else 0)
            ws.cell(row=row_num, column=8, value=medicine.expiry_date.strftime('%Y-%m-%d') if medicine.expiry_date else 'N/A')
            ws.cell(row=row_num, column=9, value='In Stock' if medicine.stock_quantity > 0 else 'Out of Stock')
        
        # Adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    elif report_type == 'daily_appointments':
        today = timezone.now().date()
        appointments = Appointment.objects.filter(appointment_date=today).select_related('patient__user', 'doctor__user')
        
        # Add title
        ws['A1'] = 'DAILY APPOINTMENTS REPORT'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f'Date: {today.strftime("%Y-%m-%d")}'
        
        # Headers
        headers = ['Time', 'Patient ID', 'Patient Name', 'Doctor', 'Type', 'Status', 'Symptoms']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row_num, apt in enumerate(appointments, 5):
            ws.cell(row=row_num, column=1, value=apt.appointment_time.strftime('%H:%M'))
            ws.cell(row=row_num, column=2, value=apt.patient.patient_id)
            ws.cell(row=row_num, column=3, value=apt.patient.user.get_full_name())
            ws.cell(row=row_num, column=4, value=apt.doctor.user.get_full_name())
            ws.cell(row=row_num, column=5, value=apt.appointment_type)
            ws.cell(row=row_num, column=6, value=apt.get_status_display())
            ws.cell(row=row_num, column=7, value=apt.symptoms[:100])
    
    wb.save(response)
    return response


@login_required
def export_report_csv(request, report_type):
    """Export report as CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.csv"'
    
    writer = csv.writer(response)
    
    # Add headers and data based on report type
    if report_type == 'medicine_stock':
        medicines = Medicine.objects.all()
        writer.writerow(['ID', 'Name', 'Category', 'Manufacturer', 'Stock', 
                        'Purchase Price', 'Selling Price', 'Expiry', 'Status'])
        
        for medicine in medicines:
            writer.writerow([
                medicine.medicine_id,
                medicine.name,
                medicine.category,
                medicine.manufacturer or 'N/A',
                medicine.stock_quantity,
                float(medicine.purchase_price) if medicine.purchase_price else 0,
                float(medicine.selling_price) if medicine.selling_price else 0,
                medicine.expiry_date.strftime('%Y-%m-%d') if medicine.expiry_date else 'N/A',
                'In Stock' if medicine.stock_quantity > 0 else 'Out of Stock'
            ])
    
    return responseh